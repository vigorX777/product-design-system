#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import subprocess
from pathlib import Path

from openpyxl import load_workbook


DOCX_SUFFIXES = {".docx"}
XLSX_SUFFIXES = {".xlsx", ".xls"}
SKIP_DIRS = {".git", ".obsidian", "node_modules"}


def clean_stem(path: Path) -> str:
    name = path.name
    while True:
        lower = name.lower()
        changed = False
        for suffix in [".docx", ".xlsx", ".xls"]:
            if lower.endswith(suffix):
                name = name[: -len(suffix)]
                changed = True
                break
        if not changed:
            break
    return name


def iter_office_files(root: Path) -> list[Path]:
    files: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        if any(part in SKIP_DIRS for part in path.parts):
            continue
        if path.suffix.lower() in DOCX_SUFFIXES | XLSX_SUFFIXES:
            files.append(path)
    return sorted(files)


def convert_docx(path: Path, dry_run: bool) -> str:
    stem = clean_stem(path)
    output_dir = path.parent / stem
    output_md = output_dir / f"{stem}.md"

    cmd = [
        "pandoc",
        str(path),
        "-t",
        "gfm",
        "--wrap=none",
        f"--extract-media={output_dir}",
        "-o",
        str(output_md),
    ]

    if dry_run:
        return f"would convert docx -> {output_md}"

    output_dir.mkdir(parents=True, exist_ok=True)
    subprocess.run(cmd, check=True)
    fix_docx_media_links(output_md)
    return f"converted docx -> {output_md}"


def fix_docx_media_links(output_md: Path) -> None:
    text = output_md.read_text(encoding="utf-8")
    text = text.replace(
        f'src="{output_md.parent.as_posix()}/media/',
        'src="./media/',
    )
    text = text.replace(
        f'(./{output_md.parent.as_posix()}/media/',
        '(./media/',
    )
    text = text.replace(
        f'(media/',
        '(./media/',
    )
    output_md.write_text(text, encoding="utf-8")


def escape_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("\n", "<br>")
    text = text.replace("|", "\\|")
    return text.strip()


def normalize_rows(rows: list[list[str]]) -> list[list[str]]:
    normalized: list[list[str]] = []
    for row in rows:
        trimmed = list(row)
        while trimmed and trimmed[-1] == "":
            trimmed.pop()
        if any(cell != "" for cell in trimmed):
            normalized.append(trimmed)

    if not normalized:
        return []

    width = max(len(row) for row in normalized)
    return [row + [""] * (width - len(row)) for row in normalized]


def sheet_to_markdown(sheet) -> str:
    rows = [[escape_cell(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    rows = normalize_rows(rows)

    if not rows:
        return "_空工作表_"

    intro_rows: list[list[str]] = []
    table_start = 0
    for idx, row in enumerate(rows):
        non_empty = [cell for cell in row if cell != ""]
        if len(non_empty) <= 1:
            intro_rows.append(row)
            table_start = idx + 1
            continue
        break

    table_rows = rows[table_start:]
    intro_lines = [f"> {next((cell for cell in row if cell), '')}" for row in intro_rows if any(row)]

    if not table_rows:
        return "\n".join(intro_lines) if intro_lines else "_空工作表_"

    if len(table_rows) == 1:
        single = "\n".join(f"- {cell}" for cell in table_rows[0] if cell)
        return "\n".join(intro_lines + ([single] if single else []))

    header = table_rows[0]
    body = table_rows[1:]
    separator = ["---"] * len(header)
    table_lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(separator) + " |",
    ]
    for row in body:
        table_lines.append("| " + " | ".join(row) + " |")
    if intro_lines:
        return "\n".join(intro_lines + [""] + table_lines)
    return "\n".join(table_lines)


def convert_xlsx(path: Path, dry_run: bool) -> str:
    stem = clean_stem(path)
    output_md = path.parent / f"{stem}.md"

    if dry_run:
        return f"would convert xlsx -> {output_md}"

    workbook = load_workbook(path, read_only=True, data_only=False)
    lines = [
        f"# {stem}",
        "",
        f"- **源文件**：`{path.name}`",
        f"- **工作表数量**：{len(workbook.worksheets)}",
        "",
    ]

    for index, sheet in enumerate(workbook.worksheets, start=1):
        lines.append(f"## {index}. {sheet.title}")
        lines.append("")
        lines.append(sheet_to_markdown(sheet))
        lines.append("")

    output_md.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return f"converted xlsx -> {output_md}"


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert product docx/xlsx files to markdown")
    parser.add_argument(
        "--root",
        default="附件/01-产品资料",
        help="Root directory to scan for Office documents",
    )
    parser.add_argument("--dry-run", action="store_true", help="Print planned actions only")
    args = parser.parse_args()

    root = Path(args.root)
    if not root.exists():
        raise SystemExit(f"root directory not found: {root}")

    files = iter_office_files(root)
    if not files:
        print("no office files found")
        return

    for path in files:
        if path.suffix.lower() in DOCX_SUFFIXES:
            print(convert_docx(path, args.dry_run))
        elif path.suffix.lower() in XLSX_SUFFIXES:
            print(convert_xlsx(path, args.dry_run))


if __name__ == "__main__":
    main()
